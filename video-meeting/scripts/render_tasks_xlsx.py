#!/usr/bin/env python3
"""
Renderer — tasks.xlsx from meeting_record.json (runs in render-env: openpyxl).

The Excel is a VIEW of the record (the single source of truth), never authored by
hand. Action items are split visually into explicit (stated/agreed) vs ai_suggested
(inferred), with decisions and open questions on their own sheets.

Usage:
  $RENDER_PY render_tasks_xlsx.py --record meeting_record.json --out tasks.xlsx
"""
import argparse
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from templates import load_template_yaml, xlsx_style  # noqa: E402

WRAP = Alignment(vertical="top", wrap_text=True)


def build_styles(style):
    """Build openpyxl style objects from an xlsx_style() dict."""
    return {
        "header_fill": PatternFill("solid", fgColor=style["header_fill"]),
        "header_font": Font(bold=True, color=style["header_font_color"],
                            name=style["font_name"]),
        "explicit_fill": PatternFill("solid", fgColor=style["explicit_fill"]),
        "suggested_fill": PatternFill("solid", fgColor=style["suggested_fill"]),
        "body_font": Font(name=style["font_name"]),
    }

COLS = [("Title", 50), ("Type", 14), ("Assignee", 20),
        ("Priority", 12), ("Source TS", 16), ("Confidence", 12)]


def style_header(ws, ncols, styles):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"


def add_items_sheet(wb, items, styles):
    ws = wb.active
    ws.title = "Action Items"
    for i, (name, width) in enumerate(COLS, 1):
        ws.cell(row=1, column=1 + (i - 1), value=name)
        ws.column_dimensions[get_column_letter(i)].width = width
    style_header(ws, len(COLS), styles)

    order = {"high": 0, "medium": 1, "low": 2}
    explicit = [t for t in items if t.get("type") == "explicit"]
    suggested = [t for t in items if t.get("type") == "ai_suggested"]
    for group in (explicit, suggested):
        group.sort(key=lambda t: order.get(t.get("priority"), 1))

    row = 2
    for t in explicit + suggested:
        ts = ", ".join(t.get("source_ts", []) or [])
        vals = [t.get("title", ""), t.get("type", ""), t.get("assignee", ""),
                t.get("priority", ""), ts, t.get("confidence", "")]
        fill = styles["explicit_fill"] if t.get("type") == "explicit" else styles["suggested_fill"]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.fill = fill
            cell.alignment = WRAP
            cell.font = styles["body_font"]
        row += 1
    if row == 2:
        ws.cell(row=2, column=1, value="(no action items)")


def add_list_sheet(wb, title, values, styles):
    ws = wb.create_sheet(title)
    ws.cell(row=1, column=1, value=title)
    ws.column_dimensions["A"].width = 90
    style_header(ws, 1, styles)
    if values:
        for i, v in enumerate(values, start=2):
            cell = ws.cell(row=i, column=1, value=v)
            cell.alignment = WRAP
            cell.font = styles["body_font"]
    else:
        ws.cell(row=2, column=1, value=f"(no {title.lower()})")


def render(record_path, out_path, template_dir=None):
    with open(record_path, "r", encoding="utf-8") as fh:
        record = json.load(fh)
    data = load_template_yaml(template_dir) if template_dir else {}
    styles = build_styles(xlsx_style(data))

    wb = Workbook()
    add_items_sheet(wb, record.get("action_items", []), styles)
    add_list_sheet(wb, "Decisions", record.get("decisions", []), styles)
    add_list_sheet(wb, "Open Questions", record.get("open_questions", []), styles)

    os.makedirs(os.path.dirname(os.path.abspath(out_path)) or ".", exist_ok=True)
    wb.save(out_path)
    return len(record.get("action_items", []))


def main():
    ap = argparse.ArgumentParser(description="Render tasks.xlsx from the meeting record.")
    ap.add_argument("--record", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--template-dir", default=None)
    args = ap.parse_args()

    if not os.path.isfile(args.record):
        sys.exit(f"record not found: {args.record}")
    n = render(args.record, args.out, args.template_dir)
    sys.stderr.write(f"  tasks.xlsx: {n} action items -> {args.out}\n")


if __name__ == "__main__":
    main()
