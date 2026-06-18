import os
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

try:
    import openpyxl  # noqa: F401
    import yaml  # noqa: F401
    HAVE_DEPS = True
except ImportError:
    HAVE_DEPS = False


@unittest.skipUnless(HAVE_DEPS, "openpyxl + pyyaml required")
class XlsxTheming(unittest.TestCase):
    def _record(self):
        return {
            "action_items": [
                {"title": "Do X", "type": "explicit", "priority": "high"},
                {"title": "Maybe Y", "type": "ai_suggested", "priority": "low"},
            ],
            "decisions": ["Decided A"],
            "open_questions": ["What about B?"],
        }

    def _write(self, d, record, template_dir=None):
        import json
        from render_tasks_xlsx import render
        rec = os.path.join(d, "rec.json")
        with open(rec, "w") as fh:
            json.dump(record, fh)
        out = os.path.join(d, "tasks.xlsx")
        render(rec, out, template_dir)
        return out

    def test_defaults_when_no_template(self):
        from openpyxl import load_workbook
        with tempfile.TemporaryDirectory() as d:
            out = self._write(d, self._record(), None)
            wb = load_workbook(out)
            ws = wb["Action Items"]
            # header fill uses the default 1F4E78 (openpyxl stores ARGB)
            self.assertIn("1F4E78", ws.cell(row=1, column=1).fill.fgColor.rgb)

    def test_template_yaml_overrides_header_fill(self):
        from openpyxl import load_workbook
        with tempfile.TemporaryDirectory() as d:
            tdir = os.path.join(d, "tpl")
            os.mkdir(tdir)
            with open(os.path.join(tdir, "template.yaml"), "w") as fh:
                fh.write("xlsx:\n  header_fill: '00AA00'\n")
            out = self._write(d, self._record(), tdir)
            wb = load_workbook(out)
            ws = wb["Action Items"]
            self.assertIn("00AA00", ws.cell(row=1, column=1).fill.fgColor.rgb)


if __name__ == "__main__":
    unittest.main()
